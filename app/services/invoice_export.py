"""CA-facing invoice export (INV-07, D-02).

Produces a spreadsheet with **one row per invoice line**, in a documented and versioned
column set, as XLSX and CSV, built entirely from the immutable `Invoice`/`InvoiceLine`
snapshot (never from `Order`, `OrderItem`, `SKU` or `Retailer`).

Contract: `EXPORT_FORMAT_VERSION` and `INVOICE_EXPORT_COLUMNS` below are the contract. The
header row they produce is pinned by the fixture at
`app/tests/fixtures/invoice_export_v1.json`, compared exactly by
`app/tests/test_invoice_export.py`. Changing any column — adding, removing, renaming or
reordering — requires bumping `EXPORT_FORMAT_VERSION` and updating that fixture in the same
commit. The column meanings are documented in `docs/invoice-export-format.md`.

Out of scope, per D-02: a Tally importer and a GSTR-1 JSON generator. This module only
produces the CA-consumable Excel/CSV.
"""
import csv
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, selectinload

from app.models.invoice import Invoice

EXPORT_FORMAT_VERSION = "1.0"

# (header, kind) pairs. `header` is the contract — it is the literal cell text emitted by
# both writers and the value pinned in the checked-in fixture. `kind` drives XLSX
# number-formatting only; it has no bearing on the CSV writer, which always writes text.
INVOICE_EXPORT_COLUMNS = (
    ("Invoice Number", "string"),
    ("Invoice Date", "date"),
    ("Invoice Type", "string"),
    ("Reverse Charge", "string"),
    ("Supplier GSTIN", "string"),
    ("Buyer Name", "string"),
    ("Buyer GSTIN", "string"),
    ("Place Of Supply", "string"),
    ("Order ID", "integer"),
    ("Line Number", "integer"),
    ("Description", "string"),
    ("HSN", "string"),
    ("Quantity", "integer"),
    ("UQC", "string"),
    ("Unit Rate", "money"),
    ("Discount Amount", "money"),
    ("Taxable Value", "money"),
    ("CGST Rate", "rate"),
    ("CGST Amount", "money"),
    ("SGST Rate", "rate"),
    ("SGST Amount", "money"),
    ("IGST Rate", "rate"),
    ("IGST Amount", "money"),
    ("Cess Rate", "rate"),
    ("Cess Amount", "money"),
    ("Line Total", "money"),
    ("Invoice Grand Total", "money"),
)


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    return date.fromisoformat(value)


def build_invoice_export_rows(
    db: Session,
    from_date: str | None = None,
    to_date: str | None = None,
) -> list[dict]:
    """One dict per invoice line, keyed by the `INVOICE_EXPORT_COLUMNS` headers.

    Every value comes from `Invoice`/`InvoiceLine` columns only — the export inherits the
    snapshot's immutability guarantee only if the snapshot is its only source. Reads
    `Invoice.lines` via `selectinload` (no per-invoice fan-out query); the relationship is
    already ordered by `InvoiceLine.line_number`.
    """
    query = db.query(Invoice).options(selectinload(Invoice.lines))

    from_dt = _parse_date(from_date)
    to_dt = _parse_date(to_date)
    if from_dt:
        query = query.filter(Invoice.invoice_date >= datetime.combine(from_dt, datetime.min.time()))
    if to_dt:
        # Inclusive of the whole `to_date` day: less-than the start of the following day.
        query = query.filter(
            Invoice.invoice_date < datetime.combine(to_dt + timedelta(days=1), datetime.min.time())
        )

    invoices = query.order_by(Invoice.invoice_number).all()

    rows: list[dict] = []
    for invoice in invoices:
        invoice_date = invoice.invoice_date.date() if invoice.invoice_date else None
        for line in invoice.lines:
            rows.append({
                "Invoice Number": invoice.invoice_number,
                "Invoice Date": invoice_date,
                "Invoice Type": invoice.invoice_type,
                "Reverse Charge": "Y" if invoice.reverse_charge else "N",
                "Supplier GSTIN": invoice.supplier_gstin,
                "Buyer Name": invoice.buyer_name,
                "Buyer GSTIN": invoice.buyer_gstin,
                "Place Of Supply": invoice.place_of_supply,
                "Order ID": invoice.order_id,
                "Line Number": line.line_number,
                "Description": line.description,
                "HSN": line.hsn_code,
                "Quantity": line.quantity,
                "UQC": line.uqc,
                "Unit Rate": line.unit_rate,
                "Discount Amount": line.discount_amount,
                "Taxable Value": line.taxable_value,
                "CGST Rate": line.cgst_rate,
                "CGST Amount": line.cgst_amount,
                "SGST Rate": line.sgst_rate,
                "SGST Amount": line.sgst_amount,
                "IGST Rate": line.igst_rate,
                "IGST Amount": line.igst_amount,
                "Cess Rate": line.cess_rate,
                "Cess Amount": line.cess_amount,
                "Line Total": line.line_total,
                "Invoice Grand Total": invoice.grand_total,
            })
    return rows


_MONEY_OR_RATE_FORMAT = "0.00"
_INTEGER_FORMAT = "0"
_DATE_FORMAT = "yyyy-mm-dd"


def export_invoices_xlsx(
    db: Session,
    from_date: str | None = None,
    to_date: str | None = None,
) -> BytesIO:
    rows = build_invoice_export_rows(db, from_date=from_date, to_date=to_date)
    headers = [header for header, _ in INVOICE_EXPORT_COLUMNS]
    kinds = [kind for _, kind in INVOICE_EXPORT_COLUMNS]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoice Lines"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append([row.get(header) for header in headers])

    sheet.freeze_panes = "A2"

    for col_idx, kind in enumerate(kinds, start=1):
        if kind in ("money", "rate"):
            fmt = _MONEY_OR_RATE_FORMAT
        elif kind == "integer":
            fmt = _INTEGER_FORMAT
        elif kind == "date":
            fmt = _DATE_FORMAT
        else:
            continue
        for cell in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell[0].number_format = fmt

    widths = [len(header) for header in headers]
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=len(headers)):
        for idx, cell in enumerate(row):
            value = cell.value
            text = "" if value is None else str(value)
            if len(text) > widths[idx]:
                widths[idx] = len(text)
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), 40)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _csv_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def export_invoices_csv(
    db: Session,
    from_date: str | None = None,
    to_date: str | None = None,
) -> BytesIO:
    rows = build_invoice_export_rows(db, from_date=from_date, to_date=to_date)
    headers = [header for header, _ in INVOICE_EXPORT_COLUMNS]

    text_buffer = StringIO()
    writer = csv.DictWriter(text_buffer, fieldnames=headers)
    writer.writeheader()
    for row in rows:
        writer.writerow({header: _csv_cell(row.get(header)) for header in headers})

    output = BytesIO(text_buffer.getvalue().encode("utf-8"))
    output.seek(0)
    return output
