from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Sheet name → column headers. Order matters: Brands must precede SKUs.
SHEET_HEADERS: dict[str, list[str]] = {
    "Brands": [
        "Name",
        "POC Name",
        "POC Phone",
        "POC Email",
    ],
    "Retailers": [
        "Name",
        "Mobile Number",
        "Address Line 1",
        "Address Line 2",
        "City",
        "State",
        "Pincode",
        "GST Number",
        "Latitude",
        "Longitude",
        "External ID",
        "Assigned Salesman",
    ],
    "Warehouses": [
        "Name",
        "Location",
        "Address Line 1",
        "Address Line 2",
        "City",
        "State",
        "Pincode",
        "Manager Name",
    ],
    "SKUs": [
        "Name",
        "Brand Name",
        "HSN Code",
        "MRP",
        "Rate",
        "Distributor Landing Price",
        "Discount Amount",
        "Discount %",
        "SGST %",
        "CGST %",
        "IGST %",
        "Weight",
        "Length (cm)",
        "Width (cm)",
        "Height (cm)",
    ],
    "Beats": [
        "Name",
        "Warehouse Name",
    ],
}

TEMPLATE_FILENAME = "ascend_import_template.xlsx"


def _write_sheet(ws, headers: list[str]):
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="0F172A")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(len(header) + 4, 14)
    ws.freeze_panes = "A2"


def build_multi_sheet_template() -> BytesIO:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    for sheet_name, headers in SHEET_HEADERS.items():
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, headers)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
