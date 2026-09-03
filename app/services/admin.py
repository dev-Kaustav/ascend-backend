from datetime import datetime, date, timedelta
from io import BytesIO
import re
from sqlalchemy import func
from sqlalchemy.orm import Session, selectinload, load_only, aliased
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from app.models import (
    Brand,
    Warehouse,
    Employee,
    Retailer,
    Beat,
    SKU,
    SKUBatch,
    Inventory,
    InventoryTransaction,
    Order,
    OrderItem,
    OrderItemTax,
    CreditNote,
    Invoice,
    User,
    Group,
    CompanyProfile,
    PaymentQRImage,
)
from app.models.payment_qr_image import digest_image
from app.schemas.admin import (
    BrandCreate,
    WarehouseCreate,
    RetailerCreate,
    SKUCreate,
    InventoryReceiptCreate,
    GroupCreate,
    UserGroupUpdate,
    CompanyProfileUpdate,
)
from app.models.enums import TransactionType, OrderStatus, EmployeeRole, PaymentStatus
from app.services.transactions import transactional_session
from app.core.security import get_password_hash
from app.services.finance import calculate_order_outstanding
from app.services import inventory as inventory_service
from app.services.order import scoped_orders_query

def create_brand(db: Session, brand: BrandCreate):
    db_brand = Brand(**brand.dict())
    db.add(db_brand)
    db.commit()
    db.refresh(db_brand)
    return db_brand

def create_warehouse(db: Session, warehouse: WarehouseCreate):
    with transactional_session(db):
        manager = db.query(Employee).filter(Employee.id == warehouse.manager_id).first()
        if not manager:
            raise ValueError("Warehouse manager not found")
        if manager.role != EmployeeRole.WAREHOUSE_MANAGER:
            raise ValueError("Selected user is not a warehouse manager")
        db_warehouse = Warehouse(**warehouse.dict(exclude={"manager_id"}))
        db.add(db_warehouse)
        db.flush()
        manager.warehouse_id = db_warehouse.id
    db.refresh(db_warehouse)
    return db_warehouse

def create_retailer(db: Session, retailer: RetailerCreate):
    db_retailer = Retailer(**retailer.dict())
    db.add(db_retailer)
    db.commit()
    db.refresh(db_retailer)
    return db_retailer

def create_sku(db: Session, sku: SKUCreate, current_user=None):
    if current_user:
        role_value = getattr(current_user.role, "value", None) or getattr(current_user, "role", None)
        if role_value not in {"ADMIN", "WAREHOUSE_MANAGER"}:
            raise ValueError("Only admins or warehouse managers can create SKUs")
    db_sku = SKU(**sku.dict())
    db.add(db_sku)
    db.commit()
    db.refresh(db_sku)
    return db_sku

def get_company_profile(db: Session):
    profile = db.query(CompanyProfile).order_by(CompanyProfile.id.asc()).first()
    if profile:
        return profile
    profile = CompanyProfile(
        legal_name="Ascend Foods",
        invoice_prefix="ASC",
    )
    db.add(profile)
    db.commit()
    db.refresh(profile)
    return profile

def update_company_profile(db: Session, payload: CompanyProfileUpdate):
    profile = get_company_profile(db)
    data = payload.model_dump()
    prefix = (data.get("invoice_prefix") or "ASC").strip().upper()
    data["invoice_prefix"] = re.sub(r"[^A-Z0-9/-]", "", prefix) or "ASC"
    for key, value in data.items():
        setattr(profile, key, value)
    db.commit()
    db.refresh(profile)
    return profile

# Only raster formats reportlab can place in a PDF. SVG is rejected: reportlab cannot
# draw it without extra dependencies, and a QR that silently fails to render is worse
# than one refused at upload.
ALLOWED_QR_CONTENT_TYPES = {"image/png", "image/jpeg", "image/gif"}
MAX_QR_IMAGE_BYTES = 2 * 1024 * 1024


def set_payment_qr_image(db: Session, image_bytes: bytes, content_type: str, label: str | None = None):
    """Store a payment QR and point the company profile at it.

    Append-only: an identical re-upload resolves to the existing row (matched on digest)
    and the previous image is never deleted, because invoices issued against it still
    render from it.
    """
    if not image_bytes:
        raise ValueError("QR image is empty.")
    if len(image_bytes) > MAX_QR_IMAGE_BYTES:
        raise ValueError("QR image must be 2 MB or smaller.")
    normalized_type = (content_type or "").split(";")[0].strip().lower()
    if normalized_type not in ALLOWED_QR_CONTENT_TYPES:
        raise ValueError("QR image must be a PNG, JPEG or GIF.")

    digest = digest_image(image_bytes)
    image = db.query(PaymentQRImage).filter(PaymentQRImage.sha256 == digest).first()
    if image is None:
        image = PaymentQRImage(
            image_bytes=image_bytes,
            content_type=normalized_type,
            sha256=digest,
            label=label,
        )
        db.add(image)
        db.flush()

    profile = get_company_profile(db)
    profile.payment_qr_image_id = image.id
    db.commit()
    db.refresh(profile)
    return profile


def add_inventory_receipt(db: Session, receipt: InventoryReceiptCreate):
    with transactional_session(db):
        db_order = Order(
            from_entity_type="BRAND",
            from_entity_id=receipt.brand_id,
            to_entity_type="WAREHOUSE",
            to_entity_id=receipt.warehouse_id,
            status=OrderStatus.DELIVERED
        )
        db.add(db_order)
        db.flush()

        for item in receipt.items:
            db_item = OrderItem(
                order_id=db_order.id,
                sku_id=item.sku_id,
                quantity=item.quantity,
                unit_price=0,
                discount_amount=0
            )
            db.add(db_item)

            db_batch = SKUBatch(
                sku_id=item.sku_id,
                warehouse_id=receipt.warehouse_id,
                mfg_date=datetime.strptime(item.mfg_date, "%Y-%m-%d") if item.mfg_date else None,
                expiry_date=datetime.strptime(item.expiry_date, "%Y-%m-%d") if item.expiry_date else None,
                quantity_received=item.quantity,
                remaining_quantity=item.quantity
            )
            db.add(db_batch)
            db.flush()

            inventory = db.query(Inventory).filter(
                Inventory.sku_id == item.sku_id,
                Inventory.warehouse_id == receipt.warehouse_id
            ).with_for_update().first()
            if not inventory:
                inventory = Inventory(sku_id=item.sku_id, warehouse_id=receipt.warehouse_id, total_quantity=0)
                db.add(inventory)
            inventory.total_quantity += item.quantity

            transaction = InventoryTransaction(
                sku_id=item.sku_id,
                warehouse_id=receipt.warehouse_id,
                batch_id=db_batch.id,
                transaction_type=TransactionType.IN,
                quantity=item.quantity
            )
            db.add(transaction)
    return db_order

def _outgoing_orders_query(db: Session):
    return db.query(Order).filter(
        Order.from_entity_type == "WAREHOUSE",
        Order.to_entity_type == "RETAILER",
    )

def get_inventory(db: Session, limit: int = 50, offset: int = 0, warehouse_id: int | None = None):
    expiry_query = (
        db.query(
            SKUBatch.sku_id.label("sku_id"),
            SKUBatch.warehouse_id.label("warehouse_id"),
            func.min(SKUBatch.expiry_date).label("earliest_expiry"),
        )
        .filter(SKUBatch.expiry_date.isnot(None))
        .filter(SKUBatch.remaining_quantity > 0)
    )
    if warehouse_id:
        expiry_query = expiry_query.filter(SKUBatch.warehouse_id == warehouse_id)
    expiry_subquery = (
        expiry_query
        .group_by(SKUBatch.sku_id, SKUBatch.warehouse_id)
        .subquery()
    )

    # D1: expired_quantity is the same predicate the allocator refuses on
    # (app/services/inventory.py). Reusing expired_batch_criterion here rather
    # than re-deriving the date comparison keeps what the screen calls
    # "expired" identical to what allocation will not touch.
    as_of = inventory_service.current_business_date()
    expired_query = (
        db.query(
            SKUBatch.sku_id.label("sku_id"),
            SKUBatch.warehouse_id.label("warehouse_id"),
            func.sum(SKUBatch.remaining_quantity).label("expired_quantity"),
        )
        .filter(inventory_service.expired_batch_criterion(as_of))
    )
    if warehouse_id:
        expired_query = expired_query.filter(SKUBatch.warehouse_id == warehouse_id)
    expired_subquery = (
        expired_query
        .group_by(SKUBatch.sku_id, SKUBatch.warehouse_id)
        .subquery()
    )

    base_query = db.query(Inventory)
    if warehouse_id:
        base_query = base_query.filter(Inventory.warehouse_id == warehouse_id)
    total = base_query.count()
    rows = (
        base_query
        .with_entities(Inventory, expiry_subquery.c.earliest_expiry, expired_subquery.c.expired_quantity)
        .outerjoin(
            expiry_subquery,
            (expiry_subquery.c.sku_id == Inventory.sku_id)
            & (expiry_subquery.c.warehouse_id == Inventory.warehouse_id),
        )
        .outerjoin(
            expired_subquery,
            (expired_subquery.c.sku_id == Inventory.sku_id)
            & (expired_subquery.c.warehouse_id == Inventory.warehouse_id),
        )
        .order_by(Inventory.warehouse_id.asc(), Inventory.sku_id.asc())
        .limit(limit)
        .offset(offset)
        .all()
    )

    items = [
        {
            "sku_id": inventory.sku_id,
            "warehouse_id": inventory.warehouse_id,
            "total_quantity": round(float(inventory.total_quantity or 0), 2),
            "reserved_quantity": round(float(inventory.reserved_quantity or 0), 2),
            "earliest_expiry": earliest_expiry,
            "expired_quantity": int(expired_quantity or 0),
        }
        for inventory, earliest_expiry, expired_quantity in rows
    ]
    return items, total

def _parse_date(value: str, is_end: bool) -> datetime | None:
    if not value:
        return None
    try:
        parsed = date.fromisoformat(value)
    except ValueError:
        return None
    if is_end:
        return datetime.combine(parsed + timedelta(days=1), datetime.min.time())
    return datetime.combine(parsed, datetime.min.time())

def _apply_order_filters(
    query,
    status: str | None = None,
    payment_status: str | None = None,
    search: str | None = None,
    from_date: str | None = None,
    to_date: str | None = None,
    has_invoice: bool | None = None,
):
    if status:
        try:
            query = query.filter(Order.status == OrderStatus(status))
        except ValueError:
            pass
    if payment_status:
        try:
            query = query.filter(Order.payment_status == PaymentStatus(payment_status))
        except ValueError:
            pass
    if search:
        query = query.filter(Order.invoice.has(Invoice.invoice_number.ilike(f"%{search}%")))
    if has_invoice is True:
        query = query.filter(Order.invoice.has())
    if has_invoice is False:
        query = query.filter(~Order.invoice.has())

    from_dt = _parse_date(from_date, False)
    to_dt = _parse_date(to_date, True)
    if from_dt:
        query = query.filter(Order.created_at >= from_dt)
    if to_dt:
        query = query.filter(Order.created_at < to_dt)
    return query

def _totals_subqueries(db: Session):
    base_expr = (OrderItem.quantity * OrderItem.unit_price) - OrderItem.discount_amount
    base_sub = (
        db.query(
            OrderItem.order_id.label("order_id"),
            func.sum(base_expr).label("base_total"),
        )
        .group_by(OrderItem.order_id)
        .subquery()
    )
    tax_sub = (
        db.query(
            OrderItem.order_id.label("order_id"),
            func.sum(base_expr * (OrderItemTax.rate / 100.0)).label("tax_total"),
        )
        .join(OrderItemTax, OrderItemTax.order_item_id == OrderItem.id)
        .group_by(OrderItem.order_id)
        .subquery()
    )
    return base_sub, tax_sub

def get_orders_page(
    db: Session,
    current_user,
    limit: int = 50,
    offset: int = 0,
    status: str | None = None,
    payment_status: str | None = None,
    search: str | None = None,
    from_date: str | None = None,
    to_date: str | None = None,
    has_invoice: bool | None = None,
):
    query = scoped_orders_query(db, current_user)
    query = _apply_order_filters(
        query,
        status=status,
        payment_status=payment_status,
        search=search,
        from_date=from_date,
        to_date=to_date,
        has_invoice=has_invoice,
    )

    total = query.count()

    base_sub, tax_sub = _totals_subqueries(db)
    total_expr = func.coalesce(base_sub.c.base_total, 0)

    rows = (
        query.outerjoin(base_sub, base_sub.c.order_id == Order.id)
        .outerjoin(tax_sub, tax_sub.c.order_id == Order.id)
        .add_columns(total_expr.label("total_amount"))
        .order_by(Order.created_at.desc())
        .limit(limit)
        .offset(offset)
        .all()
    )

    items = []
    for order, total_amount in rows:
        setattr(order, "total_amount", round(float(total_amount or 0), 2))
        items.append(order)

    order_ids = [order.id for order in items]
    if order_ids:
        detailed_orders = (
            db.query(Order)
            .options(
                selectinload(Order.items).selectinload(OrderItem.taxes),
                selectinload(Order.payments),
                selectinload(Order.credit_notes).selectinload(CreditNote.items),
            )
            .filter(Order.id.in_(order_ids))
            .all()
        )
        detailed_map = {order.id: order for order in detailed_orders}
        for order in items:
            detailed = detailed_map.get(order.id)
            pending_amount = calculate_order_outstanding(detailed) if detailed else 0
            setattr(order, "pending_amount", float(pending_amount))
    warehouse_ids = {order.from_entity_id for order in items}
    retailer_ids = {order.to_entity_id for order in items}
    salesman_ids = {order.salesman_id for order in items if order.salesman_id}
    warehouse_map = {
        warehouse.id: warehouse.name
        for warehouse in db.query(Warehouse).filter(Warehouse.id.in_(warehouse_ids)).all()
    } if warehouse_ids else {}
    retailer_map = {
        retailer.id: retailer.name
        for retailer in db.query(Retailer).filter(Retailer.id.in_(retailer_ids)).all()
    } if retailer_ids else {}
    salesman_map = {
        employee.id: employee.name
        for employee in db.query(Employee).filter(Employee.id.in_(salesman_ids)).all()
    } if salesman_ids else {}
    for order in items:
        setattr(order, "warehouse_name", warehouse_map.get(order.from_entity_id))
        setattr(order, "retailer_name", retailer_map.get(order.to_entity_id))
        setattr(order, "salesman_name", salesman_map.get(order.salesman_id))
    return items, total

def _naive_datetime(value: datetime | None) -> datetime | None:
    if not value:
        return None
    return value.replace(tzinfo=None) if value.tzinfo else value

def _discount_percent(quantity: float, unit_price: float, discount_amount: float) -> float:
    base = max((quantity or 0) * (unit_price or 0), 0)
    if not base or not discount_amount:
        return 0.0
    percent = (discount_amount / base) * 100
    return max(0.0, min(100.0, percent))

def _line_amount(quantity: float, unit_price: float, discount_amount: float) -> float:
    base = (quantity or 0) * (unit_price or 0) - (discount_amount or 0)
    return round(max(base, 0.0), 2)

def get_order_export_rows(
    db: Session,
    status: str | None = None,
    payment_status: str | None = None,
    search: str | None = None,
    from_date: str | None = None,
    to_date: str | None = None,
    has_invoice: bool | None = None,
):
    order_salesman = aliased(Employee)
    retailer_salesman = aliased(Employee)
    query = (
        db.query(Order, OrderItem, SKU, Retailer, order_salesman, retailer_salesman)
        .join(OrderItem, OrderItem.order_id == Order.id)
        .join(SKU, SKU.id == OrderItem.sku_id)
        .join(Retailer, Retailer.id == Order.to_entity_id)
        .outerjoin(order_salesman, order_salesman.id == Order.salesman_id)
        .outerjoin(retailer_salesman, retailer_salesman.id == Retailer.assigned_salesman_id)
        .filter(Order.from_entity_type == "WAREHOUSE", Order.to_entity_type == "RETAILER")
    )
    query = _apply_order_filters(
        query,
        status=status,
        payment_status=payment_status,
        search=search,
        from_date=from_date,
        to_date=to_date,
        has_invoice=has_invoice,
    )
    query = query.order_by(Order.created_at.desc(), Order.id.desc(), OrderItem.id.asc())

    rows = []
    for order, item, sku, retailer, order_rep, retailer_rep in query.all():
        created_at = _naive_datetime(order.created_at)
        order_date = created_at.date() if created_at else None
        quantity = float(item.quantity or 0)
        unit_price = float(item.unit_price or 0)
        discount_amount = float(item.discount_amount or 0)
        discount_percent = _discount_percent(quantity, unit_price, discount_amount)
        amount = _line_amount(quantity, unit_price, discount_amount)
        salesman = order_rep or retailer_rep
        rows.append(
            {
                "Created At": created_at,
                "Order Date": order_date,
                "Order ID": order.id,
                "Customer Name ( Retailer Name)": retailer.name,
                "Retailer ID": retailer.id,
                "SKU": sku.name,
                "SKU Quantity": quantity,
                "Salesman Name": salesman.name if salesman else None,
                "Salesman Number": str(salesman.phone_number) if salesman and salesman.phone_number else None,
                "MRP": float(sku.mrp) if sku.mrp is not None else None,
                "Discount %": discount_percent,
                "Amount": amount,
                "Rate": unit_price,
            }
        )
    return rows

def export_orders_excel(
    db: Session,
    status: str | None = None,
    payment_status: str | None = None,
    search: str | None = None,
    from_date: str | None = None,
    to_date: str | None = None,
    has_invoice: bool | None = None,
) -> BytesIO:
    rows = get_order_export_rows(
        db,
        status=status,
        payment_status=payment_status,
        search=search,
        from_date=from_date,
        to_date=to_date,
        has_invoice=has_invoice,
    )

    headers = [
        "Created At",
        "Order Date",
        "Order ID",
        "Customer Name ( Retailer Name)",
        "Retailer ID",
        "SKU",
        "SKU Quantity",
        "Salesman Name",
        "Salesman Number",
        "MRP",
        "Discount %",
        "Amount",
        "Rate",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append([row.get(header) for header in headers])

    sheet.freeze_panes = "A2"

    date_formats = {1: "yyyy-mm-dd hh:mm", 2: "yyyy-mm-dd"}
    for col_idx, fmt in date_formats.items():
        for cell in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell[0].number_format = fmt

    for col_idx in (7, 10, 11, 12, 13):
        for cell in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            cell[0].number_format = "0.00"

    widths = [len(header) for header in headers]
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=len(headers)):
        for idx, cell in enumerate(row):
            value = cell.value
            text = "" if value is None else str(value)
            if len(text) > widths[idx]:
                widths[idx] = len(text)
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), 40)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def get_admin_summary(db: Session, from_date: str | None = None, to_date: str | None = None):
    from_dt = _parse_date(from_date, False)
    to_dt = _parse_date(to_date, True)

    def _apply_date(q):
        if from_dt:
            q = q.filter(Order.created_at >= from_dt)
        if to_dt:
            q = q.filter(Order.created_at < to_dt)
        return q

    base_outgoing = db.query(Order).filter(
        Order.from_entity_type == "WAREHOUSE", Order.to_entity_type == "RETAILER"
    )
    orders_count = _apply_date(base_outgoing).count()

    payment_rows = _apply_date(
        db.query(Order.payment_status, func.count(Order.id))
        .filter(Order.from_entity_type == "WAREHOUSE", Order.to_entity_type == "RETAILER")
    ).group_by(Order.payment_status).all()
    payment_status = {row[0].value if hasattr(row[0], "value") else row[0]: row[1] for row in payment_rows}

    base_sub, tax_sub = _totals_subqueries(db)
    total_expr = func.coalesce(base_sub.c.base_total, 0)

    revenue_q = (
        db.query(func.coalesce(func.sum(total_expr), 0))
        .select_from(Order)
        .outerjoin(base_sub, base_sub.c.order_id == Order.id)
        .outerjoin(tax_sub, tax_sub.c.order_id == Order.id)
        .filter(Order.from_entity_type == "WAREHOUSE", Order.to_entity_type == "RETAILER")
    )
    if from_dt:
        revenue_q = revenue_q.filter(Order.created_at >= from_dt)
    if to_dt:
        revenue_q = revenue_q.filter(Order.created_at < to_dt)
    outgoing_total = round(float(revenue_q.scalar() or 0), 2)

    outstanding_q = (
        db.query(func.coalesce(func.sum(total_expr), 0))
        .select_from(Order)
        .outerjoin(base_sub, base_sub.c.order_id == Order.id)
        .outerjoin(tax_sub, tax_sub.c.order_id == Order.id)
        .filter(
            Order.from_entity_type == "WAREHOUSE",
            Order.to_entity_type == "RETAILER",
            Order.payment_status != PaymentStatus.PAID,
        )
    )
    if from_dt:
        outstanding_q = outstanding_q.filter(Order.created_at >= from_dt)
    if to_dt:
        outstanding_q = outstanding_q.filter(Order.created_at < to_dt)
    outstanding_total = round(float(outstanding_q.scalar() or 0), 2)

    # Timeline: use date range if given, else last 30 days
    timeline_since = from_dt if from_dt else (datetime.utcnow() - timedelta(days=30))
    timeline_q = (
        db.query(func.date(Order.created_at), func.count(Order.id))
        .filter(
            Order.created_at >= timeline_since,
            Order.from_entity_type == "WAREHOUSE",
            Order.to_entity_type == "RETAILER",
        )
    )
    if to_dt:
        timeline_q = timeline_q.filter(Order.created_at < to_dt)
    timeline_rows = (
        timeline_q
        .group_by(func.date(Order.created_at))
        .order_by(func.date(Order.created_at))
        .all()
    )
    orders_over_time = [
        {"label": day.strftime("%d %b"), "orders": count}
        for day, count in timeline_rows
        if day
    ]

    warehouse_map = {w.id: w.name for w in db.query(Warehouse).all()}
    sales_q = (
        db.query(Order.from_entity_id, func.count(Order.id))
        .filter(Order.from_entity_type == "WAREHOUSE", Order.to_entity_type == "RETAILER")
    )
    if from_dt:
        sales_q = sales_q.filter(Order.created_at >= from_dt)
    if to_dt:
        sales_q = sales_q.filter(Order.created_at < to_dt)
    sales_rows = sales_q.group_by(Order.from_entity_id).all()
    sales_by_warehouse = [
        {
            "warehouse_id": warehouse_id,
            "warehouse_name": warehouse_map.get(warehouse_id) or f"WH-{warehouse_id}",
            "orders": count,
        }
        for warehouse_id, count in sales_rows
    ]

    # Inventory stats are current state — not date-filtered
    inventory_rows = (
        db.query(Inventory.warehouse_id, func.sum(Inventory.total_quantity))
        .group_by(Inventory.warehouse_id)
        .all()
    )
    inventory_by_warehouse = [
        {
            "warehouse_id": warehouse_id,
            "warehouse_name": warehouse_map.get(warehouse_id) or f"WH-{warehouse_id}",
            "units": float(units or 0),
        }
        for warehouse_id, units in inventory_rows
    ]
    quantities = [row[1] for row in inventory_rows]
    max_qty = max(quantities) if quantities else 0
    low_stock_threshold = max(10, int(max_qty * 0.1)) if max_qty else 10
    low_stock_count = (
        db.query(func.count(Inventory.sku_id))
        .filter(Inventory.total_quantity <= low_stock_threshold)
        .scalar()
        or 0
    )

    return {
        "orders": {
            "total": orders_count,
            "revenue": float(outgoing_total),
            "outstanding": float(outstanding_total),
            "payment_status": payment_status,
        },
        "inventory": {
            "inventory_value": None,
            "low_stock_count": low_stock_count,
            "low_stock_threshold": low_stock_threshold,
            "by_warehouse": inventory_by_warehouse,
        },
        "orders_over_time": orders_over_time,
        "sales_by_warehouse": sales_by_warehouse,
    }

def list_retailers(db: Session):
    return db.query(Retailer).order_by(Retailer.name.asc()).all()

def set_retailer_beat(db: Session, retailer_id: int, beat_id: int | None):
    # RPT-01 / D2: one validated write path for beat membership. The existence check on
    # beat_id is not redundant with the FK — the SQLite test harness does not enable
    # PRAGMA foreign_keys, so without this a dangling assignment would pass every test
    # in this repo and fail only in production.
    retailer = db.query(Retailer).filter(Retailer.id == retailer_id).first()
    if not retailer:
        raise ValueError("Retailer not found")
    if beat_id is not None:
        beat = db.query(Beat).filter(Beat.id == beat_id).first()
        if not beat:
            raise ValueError("Beat not found")
    retailer.beat_id = beat_id
    db.commit()
    db.refresh(retailer)
    return retailer

def list_brands(db: Session):
    return db.query(Brand).order_by(Brand.name.asc()).all()

def list_salesmen(db: Session):
    return db.query(Employee).filter(Employee.role == EmployeeRole.SALESMAN).order_by(Employee.name.asc()).all()

def list_warehouse_managers(db: Session):
    return db.query(Employee).filter(Employee.role == EmployeeRole.WAREHOUSE_MANAGER).order_by(Employee.name.asc()).all()

def list_drivers(db: Session):
    return db.query(Employee).filter(Employee.role == EmployeeRole.DRIVER).order_by(Employee.name.asc()).all()

def list_skus(db: Session):
    return (
        db.query(SKU)
        .options(
            load_only(
                SKU.id,
                SKU.name,
                SKU.code,
                SKU.brand_id,
                SKU.hsn_code,
                SKU.distributor_landing_price,
                SKU.mrp,
                SKU.discount_amount,
                SKU.discount_percent,
                SKU.rate,
                SKU.sgst_percent,
                SKU.sgst_amount,
                SKU.cgst_percent,
                SKU.cgst_amount,
                SKU.igst_percent,
                SKU.igst_amount,
                SKU.amount,
                SKU.weight,
                SKU.length_cm,
                SKU.width_cm,
                SKU.height_cm,
            )
        )
        .order_by(SKU.name.asc())
        .all()
    )

def list_warehouses(db: Session):
    return db.query(Warehouse).order_by(Warehouse.name.asc()).all()

def get_order_status_summary(db: Session, from_date: str | None = None, to_date: str | None = None):
    from_dt = _parse_date(from_date, False)
    to_dt = _parse_date(to_date, True)
    q = (
        db.query(Order.status, func.count(Order.id))
        .filter(Order.from_entity_type == "WAREHOUSE", Order.to_entity_type == "RETAILER")
    )
    if from_dt:
        q = q.filter(Order.created_at >= from_dt)
    if to_dt:
        q = q.filter(Order.created_at < to_dt)
    rows = q.group_by(Order.status).all()
    summary = {}
    for status, count in rows:
        key = status.value if hasattr(status, "value") else status
        summary[key] = count
    return summary

def list_users(db: Session):
    return db.query(User).filter(User.deleted_at.is_(None)).order_by(User.created_at.desc()).all()

def _count_active_admins(db: Session):
    return (
        db.query(User)
        .filter(
            User.role == EmployeeRole.ADMIN,
            User.is_active == True,  # noqa: E712
            User.deleted_at.is_(None),
        )
        .count()
    )

def _normalize_email(email: str):
    return email.strip().lower()

def _default_employee_name(email: str) -> str:
    local_part = email.split("@")[0]
    parts = [part for part in re.split(r"[._-]+", local_part) if part]
    if not parts:
        return local_part
    return " ".join(part.capitalize() for part in parts)

def create_user(
    db: Session,
    email: str,
    password: str,
    role: str | None,
    employee_id: int | None,
    phone_number: int,
    is_active: bool = True,
    group_id: int | None = None,
):
    normalized_email = _normalize_email(email)
    existing = db.query(User).filter(User.email == normalized_email).first()
    if existing:
        raise ValueError("User already exists")
    group = None
    if group_id:
        group = db.query(Group).filter(Group.id == group_id).first()
        if not group:
            raise ValueError("Group not found")

    resolved_role = None
    if group:
        resolved_role = group.role
    if role:
        try:
            role_value = EmployeeRole(role)
        except ValueError:
            raise ValueError("Invalid role")
        if resolved_role and role_value != resolved_role:
            raise ValueError("Group role does not match selected role")
        resolved_role = role_value
    if not resolved_role:
        raise ValueError("Role is required")

    employee = None
    if employee_id:
        employee = db.query(Employee).filter(Employee.id == employee_id).first()
        if not employee:
            raise ValueError("Employee not found")
        if employee.role != resolved_role:
            raise ValueError("Employee role does not match selected role")
        employee.phone_number = phone_number
    else:
        employee = db.query(Employee).filter(Employee.email == normalized_email).first()
        if employee:
            if employee.role != resolved_role:
                raise ValueError("Employee role does not match selected role")
            employee.phone_number = phone_number
        else:
            employee = Employee(
                name=_default_employee_name(normalized_email),
                email=normalized_email,
                role=resolved_role,
                phone_number=phone_number,
            )
            db.add(employee)
            db.flush()

    password_hash = get_password_hash(password)
    user = User(
        email=normalized_email,
        password_hash=password_hash,
        role=resolved_role,
        group_id=group.id if group else None,
        employee_id=employee.id if employee else None,
        is_active=is_active,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user

def set_user_active(db: Session, user_id: int, is_active: bool, acting_user: User):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise ValueError("User not found")
    if user.deleted_at:
        raise ValueError("User is deleted")
    if user.role == EmployeeRole.ADMIN and not is_active:
        active_admins = _count_active_admins(db)
        if active_admins <= 1:
            raise ValueError("Cannot disable the last active admin")
        if acting_user.id == user_id:
            raise ValueError("Admins cannot disable themselves")
    user.is_active = is_active
    db.commit()
    db.refresh(user)
    return user

def set_user_role(db: Session, user_id: int, role: str, acting_user: User):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise ValueError("User not found")
    if user.deleted_at:
        raise ValueError("User is deleted")
    try:
        new_role = EmployeeRole(role)
    except ValueError:
        raise ValueError("Invalid role")
    if acting_user.id == user_id and new_role != EmployeeRole.ADMIN:
        raise ValueError("Admins cannot demote themselves")
    if user.role == EmployeeRole.ADMIN and new_role != EmployeeRole.ADMIN:
        active_admins = _count_active_admins(db)
        if active_admins <= 1:
            raise ValueError("Cannot remove the last active admin role")
    if user.group_id and user.group and user.group.role != new_role:
        user.group_id = None
    user.role = new_role
    db.commit()
    db.refresh(user)
    return user

def list_groups(db: Session):
    groups = db.query(Group).order_by(Group.name.asc()).all()
    user_counts = dict(
        db.query(User.group_id, func.count(User.id))
        .filter(
            User.group_id.isnot(None),
            User.deleted_at.is_(None),
        )
        .group_by(User.group_id)
        .all()
    )
    for group in groups:
        group.user_count = user_counts.get(group.id, 0)
    return groups

def create_group(db: Session, payload: GroupCreate):
    name = payload.name.strip()
    if not name:
        raise ValueError("Group name is required")
    existing = db.query(Group).filter(func.lower(Group.name) == name.lower()).first()
    if existing:
        raise ValueError("Group with this name already exists")
    try:
        role_value = EmployeeRole(payload.role)
    except ValueError:
        raise ValueError("Invalid role")
    group = Group(name=name, role=role_value)
    db.add(group)
    db.commit()
    db.refresh(group)
    return group

def set_user_group(db: Session, user_id: int, group_id: int | None, acting_user: User):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise ValueError("User not found")
    if user.deleted_at:
        raise ValueError("User is deleted")
    if group_id is None:
        user.group_id = None
        db.commit()
        db.refresh(user)
        return user

    group = db.query(Group).filter(Group.id == group_id).first()
    if not group:
        raise ValueError("Group not found")
    new_role = group.role
    if acting_user.id == user_id and new_role != EmployeeRole.ADMIN:
        raise ValueError("Admins cannot demote themselves")
    if user.role == EmployeeRole.ADMIN and new_role != EmployeeRole.ADMIN:
        active_admins = _count_active_admins(db)
        if active_admins <= 1:
            raise ValueError("Cannot remove the last active admin role")

    user.group_id = group_id
    user.role = new_role
    db.commit()
    db.refresh(user)
    return user

def set_user_password(db: Session, user_id: int, password: str, acting_user: User):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise ValueError("User not found")
    if user.deleted_at:
        raise ValueError("User is deleted")
    if acting_user.id == user_id and user.role != EmployeeRole.ADMIN:
        raise ValueError("Only admins can reset their own password here")
    user.password_hash = get_password_hash(password)
    db.commit()
    db.refresh(user)
    return user

def soft_delete_user(db: Session, user_id: int, acting_user: User):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise ValueError("User not found")
    if user.deleted_at:
        raise ValueError("User already deleted")
    if acting_user.id == user_id:
        raise ValueError("Users cannot delete themselves")
    if user.role == EmployeeRole.ADMIN and user.is_active:
        active_admins = _count_active_admins(db)
        if active_admins <= 1:
            raise ValueError("Cannot delete the last active admin")
    user.is_active = False
    user.deleted_at = datetime.utcnow()
    db.commit()
    db.refresh(user)
    return user
