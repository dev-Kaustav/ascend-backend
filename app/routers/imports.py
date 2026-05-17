from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.core.deps import require_admin
from app.services.excel_import import ingest_workbook, ingest_daily_sales_workbook
from app.services.excel_template import build_multi_sheet_template, TEMPLATE_FILENAME

router = APIRouter()


def _load_workbook(content: bytes):
    try:
        return openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file — ensure it is a valid .xlsx workbook")


@router.post("")
async def import_all(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    content = await file.read()
    wb = _load_workbook(content)
    try:
        result = ingest_workbook(db, wb)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return result


@router.post("/daily-sales")
async def import_daily_sales(
    file: UploadFile = File(...),
    warehouse_id: int = 1,
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    content = await file.read()
    wb = _load_workbook(content)
    try:
        result = ingest_daily_sales_workbook(db, wb, warehouse_id, current_user=current_user)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return result


@router.get("/template")
def download_template(current_user=Depends(require_admin)):
    output = build_multi_sheet_template()
    headers = {"Content-Disposition": f'attachment; filename="{TEMPLATE_FILENAME}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
