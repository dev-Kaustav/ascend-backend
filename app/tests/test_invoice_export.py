"""Tests for app/services/invoice_export.py (INV-07, D-02) — the CA-facing export.

The fixture at app/tests/fixtures/invoice_export_v1.json is the contract: the generated
header row (XLSX and CSV) must equal it exactly, or the drift alarm fires.
"""
import csv
import io
import itertools
import json
from decimal import Decimal

import openpyxl
import pytest

from app.models import Brand, Employee, Inventory, SKU, SKUBatch, Retailer, User, Warehouse
from app.models.enums import EmployeeRole
from app.schemas.order import OrderCreate, OrderItemCreate, StatusUpdate
from app.core.security import create_access_token
from app.services.invoice_export import (
    EXPORT_FORMAT_VERSION,
    INVOICE_EXPORT_COLUMNS,
    export_invoices_csv,
    export_invoices_xlsx,
)
from app.services.invoice import issue_invoice_for_order
from app.services.order import create_outgoing_order, update_order_status

_seed_counter = itertools.count()

_FIXTURE_PATH = "app/tests/fixtures/invoice_export_v1.json"


def _fixture():
    with open(_FIXTURE_PATH) as f:
        return json.load(f)


def _dispatched_order(
    db,
    warehouse_state="Delhi",
    retailer_state="Delhi",
    sgst_percent=9,
    cgst_percent=9,
    igst_percent=18,
    items=None,
    invoice_date=None,
):
    """Seed a real Warehouse/Retailer/SKU(s)/batch/inventory/user, create an order and
    drive it to OUT_FOR_DELIVERY. Returns (order, invoice). Mirrors
    app/tests/test_invoice_pdf.py::_dispatched_order, extended to accept a multi-item list."""
    n = next(_seed_counter)
    if items is None:
        items = [{"quantity": 2, "unit_price": 100}]

    brand = Brand(name=f"Export Brand {n}")
    warehouse = Warehouse(name=f"Export WH {n}", location="Loc", state=warehouse_state)
    retailer = Retailer(
        name=f"Export Retailer {n}", state=retailer_state,
        address_line1="1 Retailer Lane", city="City", pincode=100001, gst_number=None,
    )
    db.add_all([brand, warehouse, retailer])
    db.commit()

    skus = []
    for idx, _ in enumerate(items):
        sku = SKU(
            name=f"Export SKU {n}-{idx}", brand_id=brand.id, hsn_code=f"1000000{idx}",
            sgst_percent=sgst_percent, cgst_percent=cgst_percent, igst_percent=igst_percent,
        )
        db.add(sku)
        skus.append(sku)
    db.commit()

    for sku in skus:
        batch = SKUBatch(sku_id=sku.id, warehouse_id=warehouse.id, quantity_received=50, remaining_quantity=50)
        db.add(batch)
        inventory = Inventory(sku_id=sku.id, warehouse_id=warehouse.id, total_quantity=50)
        db.add(inventory)
    driver = Employee(name=f"Export Driver {n}", email=f"export-driver-{n}@ascend.com", role=EmployeeRole.DRIVER)
    user = User(email=f"export-admin-{n}@ascend.com", password_hash="x", role=EmployeeRole.ADMIN)
    db.add_all([driver, user])
    db.commit()

    order = create_outgoing_order(
        db,
        OrderCreate(
            retailer_id=retailer.id,
            warehouse_id=warehouse.id,
            items=[
                OrderItemCreate(sku_id=sku.id, quantity=item["quantity"], unit_price=item["unit_price"], discount_amount=item.get("discount_amount", 0))
                for sku, item in zip(skus, items)
            ],
        ),
        user,
    )
    order.delivery_driver_id = driver.id
    db.commit()
    update_order_status(db, order.id, StatusUpdate(status="READY_TO_SHIP"), user)
    if invoice_date is not None:
        # Issue with an explicit invoice_date up front. issue_invoice_for_order is
        # idempotent (app/services/invoice.py), so the OUT_FOR_DELIVERY transition below
        # still runs inventory dispatch but does not re-issue or overwrite this date —
        # an issued Invoice is immutable, so the date could not be set after the fact.
        issue_invoice_for_order(db, order, invoice_date=invoice_date)
    update_order_status(db, order.id, StatusUpdate(status="OUT_FOR_DELIVERY"), user)
    db.refresh(order)
    return order, order.invoice


def test_generated_header_row_matches_the_checked_in_fixture(db):
    fixture = _fixture()
    assert EXPORT_FORMAT_VERSION == fixture["version"]

    output = export_invoices_xlsx(db)
    workbook = openpyxl.load_workbook(output)
    sheet = workbook.active
    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    assert header_row == fixture["columns"]


def test_csv_header_matches_the_same_fixture(db):
    fixture = _fixture()
    output = export_invoices_csv(db)
    text = output.getvalue().decode("utf-8")
    reader = csv.reader(io.StringIO(text))
    header_row = next(reader)

    assert header_row == fixture["columns"]


def test_one_row_per_invoice_line(db):
    order, invoice = _dispatched_order(
        db,
        items=[
            {"quantity": 1, "unit_price": 50},
            {"quantity": 2, "unit_price": 75},
            {"quantity": 3, "unit_price": 25},
        ],
    )

    output = export_invoices_xlsx(db)
    workbook = openpyxl.load_workbook(output)
    sheet = workbook.active
    headers = [h for h, _ in INVOICE_EXPORT_COLUMNS]
    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))

    assert len(data_rows) == 3
    line_number_idx = headers.index("Line Number")
    invoice_number_idx = headers.index("Invoice Number")
    invoice_date_idx = headers.index("Invoice Date")
    grand_total_idx = headers.index("Invoice Grand Total")

    assert [row[line_number_idx] for row in data_rows] == [1, 2, 3]
    for row in data_rows:
        assert row[invoice_number_idx] == invoice.invoice_number
        cell_date = row[invoice_date_idx]
        cell_date = cell_date.date() if hasattr(cell_date, "date") else cell_date
        assert cell_date.isoformat() == invoice.invoice_date.date().isoformat()
        assert Decimal(str(row[grand_total_idx])) == invoice.grand_total


def test_exported_values_match_the_stored_invoice(db):
    order, invoice = _dispatched_order(db)
    line = invoice.lines[0]

    output = export_invoices_xlsx(db)
    workbook = openpyxl.load_workbook(output)
    sheet = workbook.active
    headers = [h for h, _ in INVOICE_EXPORT_COLUMNS]
    row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    row_by_header = dict(zip(headers, row))

    assert Decimal(str(row_by_header["Taxable Value"])) == line.taxable_value
    assert Decimal(str(row_by_header["CGST Amount"])) == line.cgst_amount
    assert Decimal(str(row_by_header["SGST Amount"])) == line.sgst_amount
    assert Decimal(str(row_by_header["Line Total"])) == line.line_total
    assert row_by_header["Quantity"] == line.quantity
    assert row_by_header["HSN"] == line.hsn_code
    assert row_by_header["UQC"] == line.uqc
    assert Decimal(str(row_by_header["Invoice Grand Total"])) == invoice.grand_total


def test_export_is_unaffected_by_editing_the_underlying_order(db):
    order, invoice = _dispatched_order(db)
    first_output = export_invoices_csv(db).getvalue()

    item = order.items[0]
    item.unit_price = Decimal("999.00")
    retailer = db.query(Retailer).filter(Retailer.id == order.to_entity_id).first()
    retailer.name = "Renamed After Issue"
    db.commit()

    second_output = export_invoices_csv(db).getvalue()

    assert first_output == second_output


def test_inter_state_line_carries_igst_and_zero_cgst_sgst(db):
    inter_order, inter_invoice = _dispatched_order(db, warehouse_state="Delhi", retailer_state="Karnataka")
    intra_order, intra_invoice = _dispatched_order(db, warehouse_state="Delhi", retailer_state="Delhi")

    headers = [h for h, _ in INVOICE_EXPORT_COLUMNS]
    output = export_invoices_xlsx(db)
    workbook = openpyxl.load_workbook(output)
    sheet = workbook.active
    rows_by_invoice_number = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_by_header = dict(zip(headers, row))
        rows_by_invoice_number[row_by_header["Invoice Number"]] = row_by_header

    inter_row = rows_by_invoice_number[inter_invoice.invoice_number]
    assert Decimal(str(inter_row["IGST Amount"])) > 0
    assert Decimal(str(inter_row["CGST Amount"])) == 0
    assert Decimal(str(inter_row["SGST Amount"])) == 0

    intra_row = rows_by_invoice_number[intra_invoice.invoice_number]
    assert Decimal(str(intra_row["CGST Amount"])) > 0
    assert Decimal(str(intra_row["SGST Amount"])) > 0
    assert Decimal(str(intra_row["IGST Amount"])) == 0


def test_date_range_filter_selects_by_invoice_date(db):
    from datetime import datetime

    order_a, invoice_a = _dispatched_order(db, invoice_date=datetime(2026, 1, 5))
    order_b, invoice_b = _dispatched_order(db, invoice_date=datetime(2026, 6, 15))

    output = export_invoices_csv(db, from_date="2026-01-01", to_date="2026-01-31")
    text = output.getvalue().decode("utf-8")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)

    invoice_numbers = {row["Invoice Number"] for row in rows}
    assert invoice_a.invoice_number in invoice_numbers
    assert invoice_b.invoice_number not in invoice_numbers


def test_export_endpoint_returns_xlsx_and_csv_and_rejects_other_formats(client, db):
    admin = User(email="export-endpoint-admin@ascend.com", password_hash="x", role=EmployeeRole.ADMIN)
    db.add(admin)
    db.commit()
    token = create_access_token({"user_id": admin.id, "role": EmployeeRole.ADMIN.value})
    headers = {"Authorization": f"Bearer {token}"}

    xlsx_response = client.get("/admin/invoices/export?format=xlsx", headers=headers)
    assert xlsx_response.status_code == 200
    assert xlsx_response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    csv_response = client.get("/admin/invoices/export?format=csv", headers=headers)
    assert csv_response.status_code == 200
    assert csv_response.headers["content-type"].startswith("text/csv")

    bad_response = client.get("/admin/invoices/export?format=pdf", headers=headers)
    assert bad_response.status_code == 400


def test_export_of_an_empty_range_produces_a_header_only_file(db):
    fixture = _fixture()

    output = export_invoices_xlsx(db, from_date="1999-01-01", to_date="1999-01-02")
    workbook = openpyxl.load_workbook(output)
    sheet = workbook.active

    assert sheet.max_row == 1
    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header_row == fixture["columns"]
