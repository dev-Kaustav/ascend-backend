from datetime import datetime
from io import BytesIO

import openpyxl
import pytest
from openpyxl import Workbook

from app.models import Account, Beat, Invoice, Order, OrderItem, OrderTrail, Retailer
from app.services.excel_import import (
    _is_date_sheet,
    _parse_date_from_sheet_or_row,
    ingest_daily_sales_workbook,
)


def _workbook(sheets: dict[str, list[list]]):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return openpyxl.load_workbook(buffer, read_only=True, data_only=True)


@pytest.mark.parametrize("sheet_name", ["3 Feb", "3rd Feb", "3 _APR", "3FEB", "12"])
def test_date_sheet_detection_accepts_supported_names(sheet_name):
    assert _is_date_sheet(sheet_name) is True


@pytest.mark.parametrize(
    "sheet_name",
    [
        "Recovery",
        "Summary",
        "Whole Data",
        "Bank",
        "Sheet1",
        "Credit",
        "Outstanding",
        "Credit Notes",
    ],
)
def test_date_sheet_detection_rejects_non_sales_names(sheet_name):
    assert _is_date_sheet(sheet_name) is False


def test_date_parser_uses_matching_datetime_verbatim():
    row_date = datetime(2026, 2, 3, 14, 30)

    assert _parse_date_from_sheet_or_row("3 Feb", row_date) == row_date


def test_date_parser_disambiguates_slash_date_with_sheet_month():
    assert _parse_date_from_sheet_or_row("3 Feb", "2/3/2026") == datetime(2026, 2, 3)


def test_date_parser_corrects_transposed_datetime_with_sheet_month():
    assert _parse_date_from_sheet_or_row("3 Feb", datetime(2026, 3, 2)) == datetime(2026, 2, 3)


def test_date_parser_falls_back_to_hardcoded_historical_year():
    """Undated historical rows are intentionally assigned to the hardcoded year 2026."""
    assert _parse_date_from_sheet_or_row("3 Feb", None) == datetime(2026, 2, 3)


def test_daily_sales_import_is_blocked_before_creating_financial_records(db):
    """RPT-10: importing summary amounts cannot produce accurate immutable invoices."""
    workbook = _workbook(
        {
            "3 Feb": [
                [
                    "Date",
                    "User",
                    "Outlet_Name",
                    "Outlet_Id",
                    "Bill_No",
                    "Amount",
                    "Cash",
                    "Online",
                    "Cheque",
                    "Credit",
                    "Rider",
                    "Status",
                    "Define",
                    "Remark",
                ],
                [
                    None,
                    "BEAT-1",
                    "Safety Block Outlet",
                    "OUTLET-1",
                    "BILL-1",
                    1250,
                    250,
                    0,
                    0,
                    1000,
                    "Driver One",
                    "Delivered",
                    None,
                    "must not be imported",
                ],
            ]
        }
    )

    with pytest.raises(ValueError, match="RPT-10"):
        ingest_daily_sales_workbook(db, workbook, warehouse_id=1)

    for model in (Retailer, Beat, Order, OrderItem, Invoice, Account, OrderTrail):
        assert db.query(model).count() == 0
