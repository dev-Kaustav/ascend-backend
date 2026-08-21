from io import BytesIO

import openpyxl
import pytest
from openpyxl import Workbook

from app.core.security import create_access_token
from app.models import Account, Beat, Invoice, Order, OrderItem, OrderTrail, Retailer, User
from app.models.enums import EmployeeRole
from app.services.excel_import import ingest_daily_sales_workbook


def _workbook_bytes(sheets: dict[str, list[list]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _loaded_workbook(sheets: dict[str, list[list]]):
    return openpyxl.load_workbook(
        BytesIO(_workbook_bytes(sheets)),
        read_only=True,
        data_only=True,
    )


def _admin_headers(db) -> dict[str, str]:
    user = User(
        email="daily-sales-admin@example.com",
        password_hash="unused-in-token-auth",
        role=EmployeeRole.ADMIN,
    )
    db.add(user)
    db.commit()
    token = create_access_token({"user_id": user.id, "role": user.role.value})
    return {"Authorization": f"Bearer {token}"}


def _assert_no_import_rows(db):
    for model in (Retailer, Beat, Order, OrderItem, Invoice, Account, OrderTrail):
        assert db.query(model).count() == 0


def test_recovery_workbook_is_blocked_before_bill_reconciliation(db):
    """RPT-10: recovery cannot be trusted while imported invoice totals are absent."""
    workbook = _loaded_workbook(
        {
            "Recovery": [
                ["Bill No", "Amount", "Date", "Sm Name", "Status"],
                ["MISSING-BILL", 500, None, "Collector One", "Cash"],
            ]
        }
    )

    with pytest.raises(ValueError, match="RPT-10"):
        ingest_daily_sales_workbook(db, workbook, warehouse_id=1)

    _assert_no_import_rows(db)


def test_daily_sales_upload_returns_explicit_block_without_writes(client, db):
    headers = _admin_headers(db)
    content = _workbook_bytes(
        {
            "3 Feb": [
                ["Date", "User", "Outlet_Name", "Bill_No", "Amount", "Cash"],
                [None, "BEAT-1", "Blocked Outlet", "BILL-1", 1250, 250],
            ],
            "Recovery": [
                ["Bill No", "Amount", "Date", "Sm Name", "Status"],
                ["BILL-1", 500, None, "Collector One", "Cash"],
            ],
        }
    )

    response = client.post(
        "/imports/daily-sales?warehouse_id=1",
        headers=headers,
        files={
            "file": (
                "daily-sales.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert response.json()["detail"].startswith("RPT-10: daily-sales import is blocked")
    _assert_no_import_rows(db)
