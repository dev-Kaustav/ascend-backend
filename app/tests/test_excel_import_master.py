from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from openpyxl import Workbook
from sqlalchemy.exc import IntegrityError

from app.models import Beat, Brand, Employee, Retailer, SKU, Warehouse
from app.models.enums import EmployeeRole
from app.services.excel_import import ingest_workbook
from app.services.excel_template import SHEET_HEADERS


def _workbook(sheets: dict[str, list[list]]):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return openpyxl.load_workbook(buffer, read_only=True, data_only=True)


def _full_row(sheet_name: str, **values):
    return [values.get(header) for header in SHEET_HEADERS[sheet_name]]


def _employee(db, name: str, role: EmployeeRole) -> Employee:
    employee = Employee(
        name=name,
        email=f"{name.lower().replace(' ', '.')}@example.com",
        role=role,
    )
    db.add(employee)
    db.commit()
    return employee


def test_brand_import_persists_poc_fields(client, db):
    workbook = _workbook(
        {
            "Brands": [
                SHEET_HEADERS["Brands"],
                ["Crax", "Priya", 919876543210, "priya@example.com"],
            ]
        }
    )

    result = ingest_workbook(db, workbook)

    assert result["brands"] == {"inserted": 1, "errors": []}
    brand = db.query(Brand).one()
    assert brand.name == "Crax"
    assert brand.poc_name == "Priya"
    assert brand.poc_phone_number == 919876543210
    assert brand.poc_email == "priya@example.com"


def test_retailer_import_resolves_salesman_and_allows_blank_assignment(client, db):
    salesman = _employee(db, "Field Seller", EmployeeRole.SALESMAN)
    workbook = _workbook(
        {
            "Retailers": [
                SHEET_HEADERS["Retailers"],
                _full_row(
                    "Retailers",
                    Name="Assigned Outlet",
                    **{
                        "External ID": "RET-1",
                        "Assigned Salesman": salesman.name,
                        "State": "Haryana",
                        "Mobile Number": 9876543210,
                        "Pincode": 122001,
                        "Latitude": 28.46,
                        "Longitude": 77.03,
                    },
                ),
                _full_row(
                    "Retailers",
                    Name="Unassigned Outlet",
                    **{
                        "External ID": "RET-2",
                        "Assigned Salesman": None,
                        "State": "Haryana",
                    },
                ),
            ]
        }
    )

    result = ingest_workbook(db, workbook)

    assert result["retailers"] == {"inserted": 2, "errors": []}
    assigned = db.query(Retailer).filter(Retailer.external_id == "RET-1").one()
    unassigned = db.query(Retailer).filter(Retailer.external_id == "RET-2").one()
    assert assigned.assigned_salesman_id == salesman.id
    assert assigned.mobile_number == 9876543210
    assert assigned.pincode == 122001
    assert assigned.latitude == 28.46
    assert assigned.longitude == 77.03
    assert unassigned.assigned_salesman_id is None


def test_warehouse_import_links_named_manager_to_new_warehouse(client, db):
    manager = _employee(db, "Warehouse Lead", EmployeeRole.WAREHOUSE_MANAGER)
    workbook = _workbook(
        {
            "Warehouses": [
                SHEET_HEADERS["Warehouses"],
                _full_row(
                    "Warehouses",
                    Name="Central Depot",
                    **{
                        "Location": "Gurgaon",
                        "City": "Gurgaon",
                        "State": "Haryana",
                        "Pincode": 122001,
                        "Manager Name": manager.name,
                    },
                ),
            ]
        }
    )

    result = ingest_workbook(db, workbook)

    assert result["warehouses"] == {"inserted": 1, "errors": []}
    warehouse = db.query(Warehouse).one()
    db.refresh(manager)
    assert warehouse.name == "Central Depot"
    assert warehouse.location == "Gurgaon"
    assert manager.warehouse_id == warehouse.id


def test_sku_import_resolves_brand_already_in_database_and_preserves_decimal_money(client, db):
    brand = Brand(name="Existing Brand")
    db.add(brand)
    db.commit()
    workbook = _workbook(
        {
            "SKUs": [
                SHEET_HEADERS["SKUs"],
                _full_row(
                    "SKUs",
                    Name="Masala Rings",
                    **{
                        "Brand Name": brand.name,
                        "HSN Code": "1905",
                        "MRP": 20,
                        "Rate": 15,
                        "Distributor Landing Price": 12.5,
                    },
                ),
            ]
        }
    )

    result = ingest_workbook(db, workbook)

    assert result["skus"] == {"inserted": 1, "errors": []}
    sku = db.query(SKU).one()
    assert sku.brand_id == brand.id
    assert sku.hsn_code == "1905"
    assert isinstance(sku.mrp, Decimal)
    assert isinstance(sku.rate, Decimal)
    assert sku.mrp == Decimal("20.00")
    assert sku.rate == Decimal("15.00")


def test_sku_import_resolves_brand_defined_in_same_workbook(client, db):
    workbook = _workbook(
        {
            "Brands": [SHEET_HEADERS["Brands"], ["Workbook Brand", None, None, None]],
            "SKUs": [
                SHEET_HEADERS["SKUs"],
                _full_row(
                    "SKUs",
                    Name="Workbook SKU",
                    **{"Brand Name": "Workbook Brand", "MRP": 30, "Rate": 25},
                ),
            ],
        }
    )

    result = ingest_workbook(db, workbook)

    assert result["brands"]["inserted"] == 1
    assert result["skus"]["inserted"] == 1
    brand = db.query(Brand).filter(Brand.name == "Workbook Brand").one()
    sku = db.query(SKU).filter(SKU.name == "Workbook SKU").one()
    assert sku.brand_id == brand.id


def test_beat_import_resolves_same_workbook_warehouse_and_allows_no_warehouse(client, db):
    manager = _employee(db, "Beat Warehouse Lead", EmployeeRole.WAREHOUSE_MANAGER)
    workbook = _workbook(
        {
            "Warehouses": [
                SHEET_HEADERS["Warehouses"],
                _full_row(
                    "Warehouses",
                    Name="Beat Depot",
                    **{
                        "State": "Haryana",
                        "Manager Name": manager.name,
                    },
                ),
            ],
            "Beats": [
                SHEET_HEADERS["Beats"],
                ["Depot Route", "Beat Depot"],
                ["Floating Route", None],
            ],
        }
    )

    result = ingest_workbook(db, workbook)

    assert result["warehouses"]["inserted"] == 1
    assert result["beats"] == {"inserted": 2, "errors": []}
    warehouse = db.query(Warehouse).filter(Warehouse.name == "Beat Depot").one()
    depot_route = db.query(Beat).filter(Beat.name == "Depot Route").one()
    floating_route = db.query(Beat).filter(Beat.name == "Floating Route").one()
    assert depot_route.warehouse_id == warehouse.id
    assert floating_route.warehouse_id is None


def test_partial_workbook_and_case_insensitive_sheet_name_report_zero_for_missing_sheets(
    client,
    db,
):
    workbook = _workbook(
        {"bRaNdS": [SHEET_HEADERS["Brands"], ["Only Brand", None, None, None]]}
    )

    result = ingest_workbook(db, workbook)

    assert result == {
        "brands": {"inserted": 1, "errors": []},
        "retailers": {"inserted": 0, "errors": []},
        "warehouses": {"inserted": 0, "errors": []},
        "skus": {"inserted": 0, "errors": []},
        "beats": {"inserted": 0, "errors": []},
    }
    assert db.query(Brand).count() == 1
    assert db.query(Retailer).count() == 0
    assert db.query(Warehouse).count() == 0
    assert db.query(SKU).count() == 0
    assert db.query(Beat).count() == 0


def test_reader_skips_banner_and_fully_blank_rows(client, db):
    workbook = _workbook(
        {
            "Brands": [
                ["Master Data Upload", None, None, None],
                SHEET_HEADERS["Brands"],
                ["First Brand", None, None, None],
                [None, None, None, None],
                ["Second Brand", None, None, None],
            ]
        }
    )

    result = ingest_workbook(db, workbook)

    assert result["brands"] == {"inserted": 2, "errors": []}
    assert [brand.name for brand in db.query(Brand).order_by(Brand.id).all()] == [
        "First Brand",
        "Second Brand",
    ]


def test_brand_validation_reports_human_row_number_for_missing_name(client, db):
    workbook = _workbook(
        {"Brands": [SHEET_HEADERS["Brands"], [None, "Priya", 1234, "p@example.com"]]}
    )

    result = ingest_workbook(db, workbook)

    assert result["brands"] == {
        "inserted": 0,
        "errors": [{"row": 2, "error": "Name is required"}],
    }
    assert db.query(Brand).count() == 0


def test_retailer_validation_reports_unknown_salesman_after_banner_offset(client, db):
    workbook = _workbook(
        {
            "Retailers": [
                ["Retailer Upload", None, None],
                SHEET_HEADERS["Retailers"],
                _full_row(
                    "Retailers",
                    Name="Bad Assignment",
                    **{
                        "External ID": "BAD-RET",
                        "Assigned Salesman": "Missing Seller",
                    },
                ),
            ]
        }
    )

    result = ingest_workbook(db, workbook)

    assert result["retailers"] == {
        "inserted": 0,
        "errors": [{"row": 3, "error": "Salesman 'Missing Seller' not found"}],
    }
    assert db.query(Retailer).count() == 0


def test_warehouse_validation_reports_missing_and_unknown_manager_rows(client, db):
    workbook = _workbook(
        {
            "Warehouses": [
                SHEET_HEADERS["Warehouses"],
                _full_row("Warehouses", Name="No Manager"),
                _full_row(
                    "Warehouses",
                    Name="Unknown Manager",
                    **{"Manager Name": "Nobody"},
                ),
            ]
        }
    )

    result = ingest_workbook(db, workbook)

    assert result["warehouses"] == {
        "inserted": 0,
        "errors": [
            {"row": 2, "error": "Manager Name is required"},
            {"row": 3, "error": "Warehouse manager 'Nobody' not found"},
        ],
    }
    assert db.query(Warehouse).count() == 0


def test_sku_validation_reports_missing_and_unknown_brand_rows(client, db):
    workbook = _workbook(
        {
            "SKUs": [
                SHEET_HEADERS["SKUs"],
                _full_row("SKUs", Name="No Brand"),
                _full_row(
                    "SKUs",
                    Name="Unknown Brand SKU",
                    **{"Brand Name": "Nobody Brand"},
                ),
            ]
        }
    )

    result = ingest_workbook(db, workbook)

    assert result["skus"] == {
        "inserted": 0,
        "errors": [
            {"row": 2, "error": "Brand Name is required"},
            {"row": 3, "error": "Brand 'Nobody Brand' not found"},
        ],
    }
    assert db.query(SKU).count() == 0


def test_beat_validation_reports_unknown_warehouse_row(client, db):
    workbook = _workbook(
        {"Beats": [SHEET_HEADERS["Beats"], ["Unknown Depot Route", "Missing Depot"]]}
    )

    result = ingest_workbook(db, workbook)

    assert result["beats"] == {
        "inserted": 0,
        "errors": [{"row": 2, "error": "Warehouse 'Missing Depot' not found"}],
    }
    assert db.query(Beat).count() == 0


def test_one_bad_row_prevents_writes_across_every_sheet(client, db):
    manager = _employee(db, "Atomicity Manager", EmployeeRole.WAREHOUSE_MANAGER)
    workbook = _workbook(
        {
            "Brands": [SHEET_HEADERS["Brands"], ["Atomic Brand", None, None, None]],
            "Retailers": [
                SHEET_HEADERS["Retailers"],
                _full_row(
                    "Retailers",
                    Name="Invalid Outlet",
                    **{"Assigned Salesman": "Missing Seller", "External ID": "ATOMIC-R"},
                ),
            ],
            "Warehouses": [
                SHEET_HEADERS["Warehouses"],
                _full_row(
                    "Warehouses",
                    Name="Atomic Depot",
                    **{"Manager Name": manager.name, "State": "Haryana"},
                ),
            ],
            "SKUs": [
                SHEET_HEADERS["SKUs"],
                _full_row(
                    "SKUs",
                    Name="Atomic SKU",
                    **{"Brand Name": "Atomic Brand", "MRP": 20},
                ),
            ],
            "Beats": [SHEET_HEADERS["Beats"], ["Atomic Route", "Atomic Depot"]],
        }
    )

    result = ingest_workbook(db, workbook)

    assert {entity: value["inserted"] for entity, value in result.items()} == {
        "brands": 0,
        "retailers": 0,
        "warehouses": 0,
        "skus": 0,
        "beats": 0,
    }
    assert result["retailers"]["errors"] == [
        {"row": 2, "error": "Salesman 'Missing Seller' not found"}
    ]
    for entity in ("brands", "warehouses", "skus", "beats"):
        assert result[entity]["errors"] == []
    assert db.query(Brand).count() == 0
    assert db.query(Retailer).count() == 0
    assert db.query(Warehouse).count() == 0
    assert db.query(SKU).count() == 0
    assert db.query(Beat).count() == 0
    db.refresh(manager)
    assert manager.warehouse_id is None


def test_duplicate_beat_rows_escape_as_a_raw_database_error_todo_rpt_09(client, db):
    """RPT-09: duplicate handling depends on incidental model constraints.

    RPT-04 requires duplicate rows to be handled, but this importer detects none.
    Unique sheets abort and imports.py returns the raw SQLAlchemy text in HTTP 400;
    unconstrained sheets double-insert. D5 leaves reject, ignore or upsert semantics
    for an explicit product decision.
    """
    workbook = _workbook(
        {"Beats": [SHEET_HEADERS["Beats"], ["Duplicate Route", None], ["Duplicate Route", None]]}
    )

    with pytest.raises(IntegrityError, match="UNIQUE constraint failed: beats.name"):
        ingest_workbook(db, workbook)

    assert db.query(Beat).count() == 0


def test_duplicate_retailer_rows_escape_as_a_raw_database_error_todo_rpt_09(client, db):
    """RPT-09 also exposes the raw unique failure for repeated retailer external IDs."""
    duplicate = _full_row(
        "Retailers",
        Name="Duplicate Outlet",
        **{"External ID": "DUPLICATE-RET", "State": "Haryana"},
    )
    workbook = _workbook(
        {"Retailers": [SHEET_HEADERS["Retailers"], duplicate, duplicate]}
    )

    with pytest.raises(
        IntegrityError,
        match="UNIQUE constraint failed: retailers.external_id",
    ):
        ingest_workbook(db, workbook)

    assert db.query(Retailer).count() == 0


def test_duplicate_brand_rows_silently_insert_twice_todo_rpt_09(client, db):
    """RPT-09: without a model constraint, identical rows are accepted twice."""
    workbook = _workbook(
        {
            "Brands": [
                SHEET_HEADERS["Brands"],
                ["Repeated Brand", None, None, None],
                ["Repeated Brand", None, None, None],
            ]
        }
    )

    result = ingest_workbook(db, workbook)

    assert result["brands"] == {"inserted": 2, "errors": []}
    assert db.query(Brand).filter(Brand.name == "Repeated Brand").count() == 2


def test_non_numeric_price_is_silently_nulled_rather_than_rejected_todo_rpt_12(client, db):
    """RPT-12: invalid numeric text becomes an unpriced catalogue row without warning.

    RPT-04 requires invalid types to be caught, but _int and _float swallow conversion
    errors. D5 records the null MRP and rate instead of changing importer semantics.
    """
    brand = Brand(name="Price Brand")
    db.add(brand)
    db.commit()
    workbook = _workbook(
        {
            "SKUs": [
                SHEET_HEADERS["SKUs"],
                _full_row(
                    "SKUs",
                    Name="Unpriced SKU",
                    **{"Brand Name": brand.name, "MRP": "n/a", "Rate": None},
                ),
            ]
        }
    )

    result = ingest_workbook(db, workbook)

    assert result["skus"] == {"inserted": 1, "errors": []}
    sku = db.query(SKU).filter(SKU.name == "Unpriced SKU").one()
    assert sku.mrp is None
    assert sku.rate is None
